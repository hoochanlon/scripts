import os
from datetime import datetime
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from snownlp import SnowNLP
import jieba
from collections import Counter
import jieba.posseg as pseg
import json
import urllib.request

def get_formatted_time():
    """
    获取格式化后的当前时间
    :return: 格式化后的当前时间字符串
    """
    now = datetime.now()
    return now.strftime('%Y-%m-%d')


def create_dir_if_not_exists(directory: str):
    """
    如果目录不存在，则创建它
    :param directory: 目录路径
    """
    os.makedirs(directory, exist_ok=True)


def get_news_from_url(url: str):
    """
    从指定的 URL 抓取热搜新闻
    :param url: 网页 URL
    :return: 热搜新闻列表
    """
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                      '(KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}
    r = requests.get(url, headers=headers)
    r.encoding = 'utf-8'
    soup = BeautifulSoup(r.text, 'html.parser')
    toutiao_resoubang = soup.find_all('div', class_='single-entry tindent')

    resoubang_list = []
    for item in toutiao_resoubang:
        spans = item.find_all('span')
        for span in spans:
            resoubang_list.append(span.string)
    return resoubang_list

# 删除空行
def delete_empty_rows(sheet_name: str, wb: Workbook):
    """
    删除指定工作表中的空行
    :param sheet_name: 工作表名称
    :param wb: Excel 工作簿对象
    """
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        if all(cell.value is None for cell in row):
            ws.delete_rows(row[0].row)

# 计数平均指数、情感得分
def calculate_average_index_and_sentiment_score(sheet_name: str, wb: Workbook):
    """
    计算指定工作表中热搜新闻的平均指数和情感得分
    :param sheet_name: 工作表名称
    :param wb: Excel 工作簿对象
    :return: 平均指数、情感得分元组
    """
    ws = wb[sheet_name]
    total_index = 0
    count = 0
    sentiment_score_list = []
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
        news_str = ''
        for cell in row:
            if cell.value is not None:
                news_str += str(cell.value)
        s = SnowNLP(news_str)
        sentiment_score = s.sentiments
        sentiment_score_list.append(sentiment_score)
        total_index += row[2].value
        count += 1
        ws.cell(row=row[0].row, column=4, value=sentiment_score)
    return (total_index / count, sum(sentiment_score_list) / len(sentiment_score_list))

# 统计词频
def calculate_word_count(sheet_names: list, wb: Workbook, stopwords_file: str):
    """
    计算工作表中出现最多的20个单词，将结果写入新的工作表中
    :param sheet_names: 工作表名称
    :param wb: Excel 工作簿对象
    :param stopwords_file: 停用词文件路径
    """
    # 请求停用词库
    response = requests.get(stopwords_file)
    stopwords = response.content.decode('utf-8').split('\n')

    # 加载停用词库
    for word in stopwords:
        jieba.del_word(word.strip())

    # 遍历所有工作表，统计词频，由于语料库词汇功能欠佳，只能粗略统计
    word_count = Counter()
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
            news_str = ''
            for cell in row:
                if cell.value is not None:
                    news_str += str(cell.value)
            words = jieba.lcut(news_str)

            # 只要是数值类型的忽略。
            # words = [word for word in words if not(word.isdigit() or (word.replace('w', '').replace('.', '').isdigit()))]
            new_words = []
            for word in words:
                 # 忽略长度为 0 或 1 的字符串
                if len(word) <= 1:
                    continue
                if not(word.isdigit() or (word.replace('w', '').replace('.', '').isdigit())):
                    new_words.append(word)
            words = new_words

            word_count.update(words)

    # 去掉停用词
    for word in list(word_count):
        if word in stopwords:
            del word_count[word]

    # 取出出现最多的50个词
    top_words = word_count.most_common(30)

    # 创建一个新的工作表
    ws = wb.create_sheet(title='词频统计')

    # 将出现最多的20个单词及其词频逐行写入新的工作表中
    ws.append(['排名', '词语', '词频'])
    for i, (word, freq) in enumerate(top_words, 1):
        ws.append([i, word, freq])


# 写入数据
def write_news_to_sheet(news_list: list, sheet_name: str, wb: Workbook):
    """
    将新闻列表写入到 Excel 工作表中
    :param news_list: 新闻列表
    :param sheet_name: 工作表名称
    :param wb: Excel 工作簿对象
    """
    ws = wb.create_sheet(title=sheet_name)

    row = []
    for i, item in enumerate(news_list, start=1):
        if i >= 156: # 只抽取50组数据（如果索引大于 156 则跳过写入操作，因为微博热榜界面有多余索引行）
            continue  
        if i % 3 == 1:
            item = item.replace("、", "")
        row.append(item)
        if i % 3 == 0:
            ws.append(row)
            row = []

    for row in ws.iter_rows(min_row=2, min_col=1):
        for cell in row:
            if cell.column == 1 or cell.column == 3:
                if isinstance(cell.value, str) and not cell.value.isnumeric():
                    # 去除字符串中的 '[置顶]' 字符，185为2、3、4排名的平均值（实时）
                    cell.value = cell.value.replace('[置顶]', '185w') 
                if isinstance(cell.value, str) and cell.value.isnumeric():
                    cell.value = int(cell.value)
                elif isinstance(cell.value, str):
                    cell.value = float(cell.value.replace('w', ''))

    ws.cell(row=1, column=3, value='指数（万）')
    ws.cell(row=1, column=4, value='情感得分')
    

def write_category_to_sheet(sheet_name: str, wb: Workbook):
    """
    将新闻事件的关键词分类信息写入到 Excel 工作表中的第五列中
    :param sheet_name: 工作表名称
    :param wb: Excel 工作簿对象
    """

    # 调用在线分类字典 json
    # 从URL获取JSON数据
    response = urllib.request.urlopen('https://ghproxy.com/https://raw.githubusercontent.com/hoochanlon/ihs-simple/main/d-json/category_news.json')
    json_data = response.read().decode('utf-8')

    # 解析JSON数据
    category_keywords = json.loads(json_data)

    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=4):
        title_str = ''
        for cell in row:
            if cell.value is not None:
                title_str += str(cell.value)
        words = pseg.cut(title_str)
        category = ''
        for word, flag in words:
            if flag.startswith('n'):
                for key, keywords in category_keywords.items():
                    if word in keywords:
                        category = key
                        break
                if category:
                    break
        if not category:
            category = '其他'
        ws.cell(row=row[0].row, column=5, value=category)
    ws.cell(row=1, column=5, value='分类')

def main():
    default_dir = os.path.join(os.path.expanduser("~"), "Desktop")
    save_path_xlsx_file = os.path.join(default_dir,
                                       "resoubang_{}.xlsx".format(get_formatted_time()))
    urls = ['http://resou.today/art/11.html', 'http://resou.today/art/22.html','http://resou.today/art/6.html']
    sheet_names = ['今日头条热榜', '抖音时事热榜', '微博热搜']

    wb = Workbook()

    for url, sheet_name in zip(urls, sheet_names):
        news_list = get_news_from_url(url)
        write_news_to_sheet(news_list, sheet_name, wb)
        delete_empty_rows(sheet_name, wb)
        
        write_category_to_sheet(sheet_name, wb)
        average_index, sentiment_score = calculate_average_index_and_sentiment_score(sheet_name, wb)
        print(f'{sheet_name} 平均指数:{average_index:.2f} 情感得分: {sentiment_score:.2f}')
       
    
    # 删除空表
    default_sheet = wb['Sheet'];wb.remove(default_sheet)

    # 停用词文件
    stopwords_file = 'https://ghproxy.com/https://raw.githubusercontent.com/goto456/stopwords/master/cn_stopwords.txt'
    # 词频统计
    calculate_word_count(sheet_names, wb, stopwords_file)

    wb.save(save_path_xlsx_file)


if __name__ == '__main__':
    main()
