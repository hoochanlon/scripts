import requests
from openpyxl import Workbook
import os
import re
from datetime import datetime

class Tools:
    # 初始化解析
    @staticmethod
    def init_res(zhihu_url):

        question_id_match = re.findall(r'question/(\d+)', zhihu_url)
        question_id = question_id_match[0]
        url = f'https://www.zhihu.com/api/v4/questions/{question_id}/feeds'

        headers = {
            'Host': 'www.zhihu.com',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.114 Safari/537.36',
            'Referer': zhihu_url
        }
        res = requests.get(url, headers=headers).json()
        print(res)
        return res, question_id

    @staticmethod
    def answer_url(url, question_id):
       answers_id = re.findall(r'answers/(\d+)', url)[0]

        # https://www.zhihu.com/question/562083816/answer/3054738773
       return 'https://www.zhihu.com/question/{}/answer/{}'.format(question_id, answers_id)


    # 格式化时间戳
    @staticmethod
    def format_timestamp(timestamp):
        dt_object = datetime.fromtimestamp(int(timestamp))
        formatted_time = dt_object.strftime("%Y-%m-%d %H:%M:%S")
        return formatted_time

    # 获取保存路径及格式
    @staticmethod
    def get_save():
        return os.path.join(os.path.join(os.path.expanduser("~"), "Desktop"),
                            "知乎问答_{}.xlsx".format(datetime.now().strftime('%Y-%m-%d')))


class ZhihuExcel:

    # 构建初始问答信息表
    @staticmethod
    def init_info(workbook, res, question_id):
        # 创建新的工作表
        worksheet = workbook.create_sheet(title='问题回答概况')

        # 写入表头
        # 标题列表
        titles = [
            '回答摘要', '点赞数', '感谢数', '评论数', '发布时间',
            '修改时间', '链接', '用户名', '用户标识'
            ]
        # 写入标题
        for col, title in enumerate(titles, start=1):
            worksheet.cell(row=1, column=col, value=title)

        data = res['data']
        row = 2  # 数据从第二行开始，因为第一行是表头
        column = 1  # 初始列数为1

        for result in data:
            worksheet.cell(row=row, column=column, value=result['target']['excerpt']);column += 1  # 回答摘要
            worksheet.cell(row=row, column=column, value=result['target']['voteup_count']);column += 1  # 点赞数
            worksheet.cell(row=row, column=column, value=result['target']['thanks_count']);column += 1  # 感谢数
            worksheet.cell(row=row, column=column, value=result['target']['comment_count']);column += 1  # 评论数
            worksheet.cell(row=row, column=column, value=Tools.format_timestamp(result['target']['created_time']));column += 1  # 发布时间
            worksheet.cell(row=row, column=column, value=Tools.format_timestamp(result['target']['updated_time']));column += 1  # 修改时间
            worksheet.cell(row=row, column=column, value=Tools.answer_url(result['target']['url'], question_id));column += 1  # 链接
            worksheet.cell(row=row, column=column, value=result['target']['author']['name']);column += 1  # 用户名
            author_token = result['target']['author']['url_token']
            if author_token:
                worksheet.cell(row=row, column=column, value=author_token);column += 1
            else:
                worksheet.cell(row=row, column=column, value='N/A');column += 1 # 用户标识，如果不存在则使用 "N/A"
            row += 1  # 每次循环结束后递增行数，写入下一行数据
            column = 1  # 每次循环结束后重置列数，从第一列开始写入数据

if __name__ == '__main__':

    # 获取知乎问题链接
    while True:
        zhihu_url = input("请粘贴知乎问题链接：")
        if re.search(r'^https://www.zhihu.com/question/\d+$', zhihu_url):
            break
        else:
            print("链接格式不正确，请重新输入。")


    # 创建工作簿和工作表
    workbook = Workbook()
    # 删除默认创建的空白Sheet
    workbook.remove(workbook.active)

    # 初始化解析
    res, question_id = Tools.init_res(zhihu_url)

    # 初始化写入
    ZhihuExcel.init_info(workbook,res, question_id)

    # 保存工作簿
    workbook.save(Tools.get_save())
