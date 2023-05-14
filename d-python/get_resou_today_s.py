import os
from datetime import datetime
import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from snownlp import SnowNLP
import jieba
from collections import Counter
import jieba.posseg as pseg
import json
import urllib.request
import stanza


# 确定保存文本格式用的。
def get_save_path_xlsx_file():
    """
    获取格式化后的当前时间
    :return: 格式化后的当前时间字符串
    """

    # 进行跨平台处理保存路径
    return os.path.join(os.path.join(os.path.expanduser("~"), "Desktop"), 
                                       "resoubang_{}.xlsx".format(datetime.now().strftime('%Y-%m-%d')))

# 解析新闻函数
def get_news_from_url(url: str):
    """
    从指定的 URL 抓取热搜新闻
    :param url: 网页 URL
    :return: 热搜新闻列表
    """
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                      '(KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}
    r = requests.get(url, headers=headers)
    r.encoding = 'utf-8'
    soup = BeautifulSoup(r.text, 'html.parser')
    toutiao_resoubang = soup.find_all('div', class_='single-entry tindent')

    resoubang_list = []
    for item in toutiao_resoubang:
        spans = item.find_all('span')
        for span in spans:
            resoubang_list.append(span.string)
    return resoubang_list

# 删除空行
def delete_empty_rows(sheet_name: str, wb: Workbook):
    """
    删除指定工作表中的空行
    :param sheet_name: 工作表名称
    :param wb: Excel 工作簿对象
    :param None 关键字
    https://notes-by-yangjinjie.readthedocs.io/zh_CN/latest/python/05-modules/openpyxl.html?highlight=openpyxl
    """
    ws = wb[sheet_name]
    # 迭代行所用的内置函数 ws.iter_rows()
    for row in ws.iter_rows():
        if all(cell.value is None for cell in row):
            ws.delete_rows(row[0].row)

# 计数平均指数、情感得分
def calculate_average_index_and_sentiment_score(sheet_name: str, wb: Workbook):
    """
    计算指定工作表中热搜新闻的平均指数和情感得分
    :param sheet_name: 工作表名称
    :param wb: Excel 工作簿对象
    :return: 平均指数、情感得分元组
    """
    ws = wb[sheet_name]
    total_index = 0
    count = 0
    sentiment_score_list = []
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
        news_str = ''
        for cell in row:
            if cell.value is not None:
                news_str += str(cell.value)
        # 功能挺多，见：https://github.com/isnowfy/snownlp
        s = SnowNLP(news_str)
        sentiment_score = s.sentiments
        sentiment_score_list.append(sentiment_score)
        # row[2] 是指每一行中的第三个单元格，也就是第三列。
        total_index += row[2].value
        count += 1
        ws.cell(row=row[0].row, column=4, value=sentiment_score)
    # 每张表的平均指数与情感得分
    return (total_index / count, sum(sentiment_score_list) / len(sentiment_score_list))

# 统计词频
def calculate_word_count(sheet_names: list, wb: Workbook):
    """
    计算工作表中出现最多的20个单词，将结果写入新的工作表中
    :param sheet_names: 工作表名称
    :param wb: Excel 工作簿对象
    :param stopwords_file: 停用词文件路径
    停用词是指在自然语言中使用频率很高，
    但通常不具有实际含义或对文本分析任务没有太大帮助的单词，如“的”，“了”等。
    """
    
    # 停用词文件
    stopwords_file = 'https://ghproxy.com/https://raw.githubusercontent.com/goto456/stopwords/master/cn_stopwords.txt'
    # 请求停用词库
    response = requests.get(stopwords_file)
    stopwords = response.content.decode('utf-8').split('\n')

    # 加载停用词库
    for word in stopwords:
        jieba.del_word(word.strip())

    # 遍历所有工作表，统计词频，由于语料库词汇功能欠佳，只能粗略统计
    word_count = Counter()
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
            news_str = ''
            for cell in row:
                if cell.value is not None:
                    news_str += str(cell.value)
            words = jieba.lcut(news_str)

            # 只要是数值类型的忽略。
            # words = [word for word in words if not(word.isdigit() or (word.replace('w', '').replace('.', '').isdigit()))]
            new_words = []
            for word in words:
                 # 忽略长度为 0 或 1 的字符串
                if len(word) <= 1:
                    continue
                # 去除数值噪声
                if not(word.isdigit() or (word.replace('w', '').replace('.', '').isdigit())):
                    new_words.append(word)
            words = new_words
            # 更新词库集
            word_count.update(words)

    # 去掉停用词
    for word in list(word_count):
        if word in stopwords:
            del word_count[word]

    # 取出出现最多的30个词
    top_words = word_count.most_common(30)
    # 创建一个新的工作表
    ws = wb.create_sheet(title='词频统计')
    # 添加行
    ws.append(['排名', '词语', '词频'])
    # 从1开始计数，而非0开始排名
    for i, (word, freq) in enumerate(top_words,1):
        ws.append([i, word, freq])


# 2023.5.9 新增分类功能
def write_category_to_sheet(sheet_name: str, wb: Workbook):
    """
    将新闻事件的关键词分类信息写入到 Excel 工作表中的第五列中
    :param sheet_name: 工作表名称
    :param wb: Excel 工作簿对象
    :jieba分词：https://github.com/fxsjy/jieba
    """

    # 调用在线分类字典 json
    # 从URL获取JSON数据
    response = urllib.request.urlopen('https://ghproxy.com/https://raw.githubusercontent.com/hoochanlon/ihs-simple/main/AQUICK/category_news.json')
    json_data = response.read().decode('utf-8')

    # 解析JSON数据
    category_keywords = json.loads(json_data)

    # 从当前sheet开始
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=4):
        title_str = ''
        for cell in row:
            if cell.value is not None:
                title_str += str(cell.value)

        # 将标题字符串分词并转换为列表数组
        words = pseg.cut(title_str)
        category = ''
        for word, flag in words:
            # 内置的字符串方法，用于检查前缀指定开头
            # if flag.startswith('n'):
            # key keywords 关键字，.items() 检索键值对
                for key, keywords in category_keywords.items():
                    if word in keywords:
                        category = key
                        break
                if category:
                    break
        if not category:
            category = '其他'
        ws.cell(row=row[0].row, column=5, value=category)


# 将已解析网页排版的数据，按规则写入到xlsx，行列对称，条理分明。
def write_news_to_sheet(news_list: list, sheet_name: str, wb: Workbook):
    """
    将新闻列表写入到 Excel 工作表中
    :param news_list: 新闻列表
    :param sheet_name: 工作表名称
    :param wb: Excel 工作簿对象
    :cell.value.isnumeric() 表示当前字符串是否能表示为一个数字
    :isinstance(cell.value, str) 表示当前值是否字符串
    """
    ws = wb.create_sheet(title=sheet_name)

    row = []
    for i, item in enumerate(news_list, start=1):
        if i >= 156: # 抽取50组数据（如果索引大于 156 是关于微博的其他介绍文章）
            continue  
        if i % 3 == 1: #  # 索引从0开始，即2%3，往后是新组数据。
            item = item.replace("、", "")
        row.append(item)
        if i % 3 == 0: # 取模被整除，3列一组，被整除说明已到3列，换行。
            ws.append(row)
            row = []
    
    # 从第二行开始迭代每一行数据
    for row in ws.iter_rows(min_row=2, min_col=1):
        for cell in row:
            if cell.column == 1 or cell.column == 3:
                # 实例判断，如果是返回 True，否则返回 False。
                if isinstance(cell.value, str) and not cell.value.isnumeric():
                    # 去除字符串中的 '[置顶]' 字符，185为2、3、4排名的平均值（实时）
                    cell.value = cell.value.replace('[置顶]', '185w') 
                if isinstance(cell.value, str) and cell.value.isnumeric():
                    cell.value = int(cell.value)
                elif isinstance(cell.value, str):
                    cell.value = float(cell.value.replace('w', ''))

    ws.cell(row=1, column=3, value='指数（万）')
    ws.cell(row=1, column=4, value='情感得分')
    ws.cell(row=1, column=5, value='分类')

# 处理整合，爬取各50条热搜，并计算文本情感值、自动化分类、词频统计
def fenmenbielei():

    result_list = []  # 定义一个空列表，用于存储每个表格的平均指数和情感得分

    urls = ['http://resou.today/art/11.html', 'http://resou.today/art/22.html','http://resou.today/art/6.html']
    sheet_names = ['今日头条热榜', '抖音时事热榜', '微博热搜']

    wb = Workbook()

    # wb.remove(wb['Sheet'])
    for url, sheet_name in zip(urls, sheet_names):
        news_list = get_news_from_url(url)
        # 写入网页解析的数据到xlsx
        write_news_to_sheet(news_list, sheet_name, wb)
        # 删除操作留下的空行
        delete_empty_rows(sheet_name, wb)
        # 分类
        write_category_to_sheet(sheet_name, wb)

        # 统计平均指数、各表平均情感值
        average_index, sentiment_score = calculate_average_index_and_sentiment_score(sheet_name, wb)
        # print(f'{sheet_name} 平均指数:{average_index:.2f} 情感得分: {sentiment_score:.2f}')
        result_list.append((average_index, sentiment_score))  # 将平均指数和情感得分以元组的形式添加到 result_list 中
    
    # 词频统计
    calculate_word_count(sheet_names, wb)

    # 删除空表，并保存为指定文件
    wb.remove(wb['Sheet']);wb.save(get_save_path_xlsx_file())

    return result_list

# 5.13新增：xlsx新闻热搜词性分析
def add_special_pos_columns(sheet):
    
    '''
    赠加文本词性识别
    stanza ： https://stanfordnlp.github.io/stanza/data_objects.html
    '''

    # 初始化中文管道
    nlp = stanza.Pipeline('zh') 

    # 在表格中添加“是否存在特定词性”和“特定词性”两列
    sheet.cell(row=1, column=6, value="是否存在特定词性")
    sheet.cell(row=1, column=7, value="特定词性")

    #读取表格中的B列数据，并在第6列标记每个单元格中是否存在特定词性，在第7列输出特定词性
    for i, cell in enumerate(sheet['B'], start=1):
        if i == 1: continue # 跳过标题行
        doc = nlp(cell.value)
        flag = False
        special_pos_list = []
        for word in doc.sentences[0].words:
            if word.upos in ['ADV', 'DET', 'ADJ', 'NUM']:
                flag = True
                special_pos_list.append(f"【{word.upos}】{word.text}")        
        sheet.cell(row=i, column=6, value="是" if flag else "否")
        sheet.cell(row=i, column=7, value=", ".join(special_pos_list))

# 加载Stanza语言模型，进行中文数词、副词、形容词分析"
def load_stanza_to_sheet():

    wb = openpyxl.load_workbook(get_save_path_xlsx_file())

    # 获取Sheet1、Sheet2和Sheet3表格对象
    # sheet1 = wb['今日头条热榜']

    # 在Sheet1中添加特定词性列
    # add_special_pos_columns(sheet1)

    # 遍历工作表名称并添加特殊词性列
    for sheet_name in ['今日头条热榜', '抖音时事热榜', '微博热搜']:
        sheet = wb[sheet_name]
        add_special_pos_columns(sheet)

    #保存Excel工作簿
    wb.save(get_save_path_xlsx_file())


def main(): 
    
    result_list =  fenmenbielei()
    load_stanza_to_sheet()
   
    print("\n热搜词性分析已加载完成，现开始计算各表热搜的指数、文本情感的平均值 \n")
    # 使用 zip 函数可以将数组打包为一个迭代器
    for sheet_name, result in zip(['今日头条热榜', '抖音时事热榜', '微博热搜'], result_list):
        average_index, sentiment_score = result
        print(f'{sheet_name} 平均指数:{average_index:.2f} 情感得分: {sentiment_score:.2f}')
    
    print("\n")

# 如果当前模块是被其他模块导入的，则该条件语句下面的代码将不会被执行。
if __name__ == '__main__':
    main()
