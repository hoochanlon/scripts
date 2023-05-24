import os
import time
import json
import jieba
import jieba.analyse
import thulac
import base64
import requests
from datetime import datetime
import urllib.request
from snownlp import SnowNLP
from collections import Counter
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
 

# 分类
def category_of_title(title, category_keywords):
    """
    对标题进行分类
    :param title: str 标题文本
    :param category_keywords: dict 分类关键词字典
    :return: str 分类名称
    """
    words = set(jieba.cut(title))
    for category, keywords in category_keywords.items():
        if words & set(keywords):   # 将列表转换成集合类型
            return category
    return '其他'

# 解析每个元素，并提取出标题、链接、来源，遍历用的。
def get_element_data(element) -> dict:
    # 获取标题和链接（经测试发现只能取到链接）
    href_element = WebDriverWait(element, 10).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, 'a'))
    )
    # 通过a元素的上级元素h2，直接取到了文本
    h2_element = WebDriverWait(element, 10).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, 'h2'))
    )
    # 通过a元素的上级元素h2，直接取到了文本
    h3_element = WebDriverWait(element, 10).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, 'h3'))
    )

    href = href_element.get_attribute('href')
    title = h2_element.text
    # 获取源网站名称，仅保留冒号后面的内容
    source = h3_element.text.split('：', 1)[-1].strip()

    # # 获取时间标识（[:6]截取前六位）
    # time_str = href.split('/')[-3]
    # datetime_obj = datetime.strptime(time_str, '%Y%m%d')

    # # 将 datetime 对象格式化为指定字符串
    # time_formatted = datetime_obj.strftime('%Y/%m/%d')

    # 获取时间标识
    time_str = href.split('/')[-3]
    if '-' in time_str:  # 如果日期符号为 '-'，则按照 '%Y-%m-%d' 的日期格式进行转换
        datetime_obj = datetime.strptime(time_str, '%Y-%m')
        time_formatted = datetime_obj.strftime('%Y/%m')
    elif '/' in time_str:  # 如果日期符号为 '/'，则按照 '%Y/%m/%d' 的日期格式进行转换
        datetime_obj = datetime.strptime(time_str, '%Y/%m/%d')
        time_formatted = datetime_obj.strftime('%Y/%m/%d')
    else:
        time_str = time_str[:6]
        # 将时间标识转换为 datetime 对象
        datetime_obj = datetime.strptime(time_str, '%Y%m')
        # 将 datetime 对象格式化为指定字符串
        time_formatted = datetime_obj.strftime('%Y/%m')


    return {'title': title, 'source': source,'time': time_formatted, 'href': href}


def selenium_url_parse(url):
    # 创建 WebDriver 对象，这里使用 Chrome 浏览器
    driver = webdriver.Chrome()

    # 访问网页
    driver.get(url)

    # 在等待 "more" 元素出现之前，先将页面滚动到底部，以确保 "more" 元素已经加载
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

    # 等待 "more" 元素出现，并进行点击
    more = WebDriverWait(driver, 10).until(
        # 个预期条件对象，用于检查页面是否出现了一个可见的元素。
        EC.visibility_of_element_located((By.CSS_SELECTOR, 'div.more')))

    # 每次加载数据为20个，拿x次加载的数据内容，不过i其实是从0算起。
    x = 1
    for i in range(x):
        time.sleep(3);more.click()

    # 获取 ul中的li 元素列表
    elements = driver.find_elements(By.CSS_SELECTOR, 'ul.list>li')

    # 遍历每个元素，提取所需数据，并存储到一个列表中
    data_list = []
    for e in elements:
        data = get_element_data(e)
        data_list.append(data)

    # 关闭浏览器
    driver.quit()

    return data_list

# 确定保存文本格式用的。
def get_save_path_xlsx_file():
    """
    获取格式化后的当前时间
    :return: 格式化后的当前时间字符串
    """
    # 进行跨平台处理保存路径
    return os.path.join(os.path.join(os.path.expanduser("~"), "Desktop"), 
                                       "piyao_{}.xlsx".format(datetime.now().strftime('%Y-%m-%d')))

# 确定字符串边界
def aphla()->str:
    return base64.b64decode("aHR0cHM6Ly93d3cucGl5YW8ub3JnLmNuL2JxL2luZGV4Lmh0bQ==").decode('utf-8')
   

def baseinfo_write_to_excel(rumor_contents) -> list:

    # 获取默认sheet
    sheet = workbook.active
    # 将默认sheet的名称更改为"谣言分析"
    sheet.title = "谣言分析"

    # 写入表头
    sheet['A1'] = '谣言标题'
    sheet['B1'] = '发布来源'
    sheet['C1'] = '辟谣时间'
    sheet['D1'] = '分类' 
    sheet['E1'] = '情感值' 
    sheet['F1'] = '链接'

    # 21：58 新增自动化分类
    # 从URL获取JSON数据，调用在线分类字典 json
    response = urllib.request.urlopen('https://ghproxy.com/https://raw.githubusercontent.com/hoochanlon/ihs-simple/main/d-json/fenlei_pypt.json')
    json_data = response.read().decode('utf-8')
    # 解析JSON数据
    category_keywords = json.loads(json_data)

    # 写入数据
    for i, data in enumerate(data_list):
        sentiment = SnowNLP(data['title']).sentiments
        category = category_of_title(data['title'], category_keywords)
        sheet.cell(row=i+2, column=1, value=data['title'])
        sheet.cell(row=i+2, column=2, value=data['source'])
        sheet.cell(row=i+2, column=3, value=data['time'])
        sheet.cell(row=i+2, column=6, value=data['href'])
        sheet.cell(row=i+2, column=4, value=category)
        sheet.cell(row=i+2, column=5, value=sentiment)

    ciyu_tongji_fenxi()

    # 保存Excel文件
    workbook.save(get_save_path_xlsx_file())

# 词频统计与语法分析
def ciyu_tongji_fenxi(): 
    '''
    该部分隶属于 baseinfo_write_to_excel(rumor_contents)管理
    '''
    # 创建新的sheet
    sheet_freq = workbook.create_sheet(title='词频统计')
    # 写入表头
    sheet_freq['A1'] = '词语'
    sheet_freq['B1'] = '出现次数'
    sheet_freq['C1'] = '词性'

    # 要去掉的字词和符号
    # stop_words = ['是', '的', '了', '，', '。', '？']
    
    # 停用词文件
    stopwords_file = 'https://ghproxy.com/https://raw.githubusercontent.com/goto456/stopwords/master/cn_stopwords.txt'
    # 请求停用词库
    stopwords = requests.get(stopwords_file).content.decode('utf-8').split('\n')

    # 分词并统计词频
    word_count = Counter()
    for data in data_list:
        words = jieba.cut(data['title'])
        for word in words:
            if word not in stopwords:
                word_count[word] += 1

    # 取前 y 个词语
    y = 15
    top = word_count.most_common(y)

# 使用THULAC进行词性标注
    word_pos = []
    for word, count in top:
        tup = thu.cut(word)
        if tup:
            word_pos.append((word, tup[0][1]))

    # 写入词频统计结果
    row = 2  # 从第二行开始写入
    for i, (word, count) in enumerate(top):
        sheet_freq.cell(row=row+i, column=1, value=word)
        sheet_freq.cell(row=row+i, column=2, value=count)
        
        pos_str = ''
        for w, pos in word_pos:
            if w == word:
                # pos_str += pos + ' '
                pos_str += pos_dict.get(pos, pos) + ' '
        sheet_freq.cell(row=row+i, column=3, value=pos_str)

def main():
    global thu
    global data_list  
    global workbook
    global pos_dict 
    
    # 全局加载中文模型
    thu = thulac.thulac()
    # 全局创建一个Workbook对象
    workbook = Workbook()
    # 参考链接：https://github.com/thunlp/THULAC#词性解释
    pos_dict = {
                "n": "名词","np": "人名","ns": "地名","ni": "机构名","nz": "其他专名",
                "m": "数词","q": "量词","mq": "数量词","t": "时间词","f": "方位词","s": "处所词",
                "v": "动词","vm": "能愿动词","vd": "趋向动词","a": "形容词","d": "副词",
                "h": "前接成分", "k": "后接成分", "i": "习语", "j": "简称",
                "r": "代词","c": "连词","p": "介词","u": "助词","y": "语气助词",
                "e": "叹词","o": "拟声词","g": "语素","w": "标点","x": "其他"
    }


# 如果当前模块是被其他模块导入的，则该条件语句下面的代码将不会被执行。
if __name__ == '__main__':
    main()
    data_list = selenium_url_parse(aphla())
    baseinfo_write_to_excel(data_list)
